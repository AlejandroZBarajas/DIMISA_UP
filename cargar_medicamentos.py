import sys


# PRODUCCION:
# docker exec -u root dimisa-mysql bash -c "microdnf install -y python3 python3-pip"
# docker cp CATALOGO.xlsx dimisa-mysql:/tmp/
# docker cp cargar_medicamentos.py dimisa-mysql:/tmp/
# docker exec dimisa-mysql pip3 install openpyxl mysql-connector-python
# docker exec dimisa-mysql python3 /tmp/cargar_medicamentos.py


# Instalacion automatica de dependencias si faltan
import subprocess
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "mysql-connector-python"], stdout=subprocess.DEVNULL)

import openpyxl
import mysql.connector

# ── Configuracion ──────────────────────────────────────────────
XLSX_PATH = "/tmp/CATALOGO.xlsx"  # ruta dentro del contenedor
DB_HOST   = "dimisa-mysql"  # nombre del servicio en docker-compose
DB_PORT   = 3306
DB_NAME   = "dimisa"              
DB_USER   = "alejandroz"
DB_PASS   = "hgmaza25"
# ───────────────────────────────────────────────────────────────

print("Leyendo archivo Excel...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws = wb.active

# Leer encabezados y encontrar indices de CLAVE y DESCRIPCION
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Columnas encontradas: {headers}")

try:
    idx_clave = headers.index("CLAVE")
    idx_desc  = headers.index("DESCRIPCION")
except ValueError as e:
    print(f"ERROR: No se encontro la columna esperada: {e}")
    sys.exit(1)

print(f"Columna CLAVE en indice {idx_clave}, DESCRIPCION en indice {idx_desc}")

# Leer todas las filas (saltando encabezado)
filas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    clave = row[idx_clave]
    desc  = row[idx_desc]
    if clave and desc:  # ignorar filas vacias
        filas.append((str(clave).strip(), str(desc).strip()))

wb.close()
print(f"Total de registros a insertar: {len(filas)}")

# Conectar a MySQL e insertar
print("Conectando a la base de datos...")
try:
    conn = mysql.connector.connect(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASS
    )
    cursor = conn.cursor()

    sql = "INSERT INTO medicamentos (clave_med, descripcion) VALUES (%s, %s)"
    cursor.executemany(sql, filas)
    conn.commit()

    print(f"✓ {cursor.rowcount} registros insertados correctamente.")
    cursor.close()
    conn.close()

except mysql.connector.Error as e:
    print(f"ERROR de base de datos: {e}")
    sys.exit(1)
